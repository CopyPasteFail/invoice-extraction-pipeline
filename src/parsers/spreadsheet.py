from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from src.io.files import read_text_best_effort

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional import safety
    load_workbook = None  # type: ignore[assignment]


@dataclass
class WorkbookSnapshot:
    sheet_names: list[str] = field(default_factory=list)
    top_rows_text: list[str] = field(default_factory=list)
    full_text: str = ""


def inspect_workbook(path: Path) -> WorkbookSnapshot:
    if load_workbook is not None:
        try:
            workbook = load_workbook(filename=path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            collected_rows: list[str] = []
            for sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
                collected_rows.append(f"Sheet: {sheet_name}")
                for row in worksheet.iter_rows(min_row=1, max_row=12, values_only=True):
                    cells = [str(cell).strip() for cell in row if cell not in (None, "")]
                    if cells:
                        collected_rows.append(" | ".join(cells))
            return WorkbookSnapshot(
                sheet_names=sheet_names,
                top_rows_text=collected_rows[:25],
                full_text="\n".join(collected_rows),
            )
        except Exception:
            fallback_text = read_text_best_effort(path)
            return WorkbookSnapshot(
                sheet_names=["Sheet1"],
                top_rows_text=fallback_text.splitlines()[:25],
                full_text=fallback_text,
            )

    fallback_text = read_text_best_effort(path)
    return WorkbookSnapshot(
        sheet_names=["Sheet1"],
        top_rows_text=fallback_text.splitlines()[:25],
        full_text=fallback_text,
    )
