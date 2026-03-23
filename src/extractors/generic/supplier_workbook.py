from __future__ import annotations

from pathlib import Path

from src.extractors.base import BaseExtractor
from src.io.json_codec import JsonObject
from src.parsers.text_utils import find_date_by_labels, find_label_value, find_money_by_labels, parse_currency, parse_date, parse_float
from src.routing.probe import ProbeResult

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional import safety
    load_workbook = None  # type: ignore[assignment]


class GenericSupplierWorkbookExtractor(BaseExtractor):
    key = "generic.supplier_workbook"
    family = "supplier_workbook"

    def extract(self, input_path: Path, probe: ProbeResult) -> dict[str, object]:
        workbook = self._extract_from_workbook(input_path)
        if workbook is not None:
            workbook["extras"] = self._empty_extras()
            return workbook

        text = probe.full_text or "\n".join(probe.workbook_labels)
        return {
            "batch_number": find_label_value(text, ["Batch Number", "Batch"]),
            "invoice_date": find_date_by_labels(text, ["Invoice Date"]),
            "period_start": find_date_by_labels(text, ["Period Start"]),
            "period_end": find_date_by_labels(text, ["Period End"]),
            "supplier_name": find_label_value(text, ["Supplier Name", "Supplier"]),
            "client_name": find_label_value(text, ["Client Name", "Customer"]),
            "total_amount": find_money_by_labels(text, ["Total Amount", "Grand Total"]),
            "currency": parse_currency(find_label_value(text, ["Currency"]) or text) or "USD",
            "categories": self._parse_categories(text),
            "extras": self._empty_extras(),
        }

    def _extract_from_workbook(self, input_path: Path) -> dict[str, object] | None:
        if load_workbook is None:
            return None
        try:
            workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
        except Exception:
            return None

        summary_sheet = workbook[workbook.sheetnames[0]]
        summary_rows = [self._row_values(row) for row in summary_sheet.iter_rows(values_only=True)]
        supplier_name = next((row[0] for row in summary_rows if row and row[0]), None)
        batch_number = None
        client_name = None
        invoice_date = None
        period_start = None
        period_end = None
        currency = "USD"
        subtotals: dict[str, float] = {}

        for row in summary_rows:
            line = " | ".join(row)
            if not batch_number:
                batch_number = find_label_value(line, ["Invoice Batch No", "Batch Number", "Batch"])
            if not client_name:
                client_name = find_label_value(line, ["Bill To", "Client Name", "Customer"])
            if not invoice_date:
                invoice_date = find_date_by_labels(line, ["Invoice Date"])
            if "Billing Period" in line and (period_start is None or period_end is None):
                period_start, period_end = self._parse_period_range(line)
                currency = parse_currency(line) or currency
            if len(row) >= 3 and row[0] in {"Storage Charges", "Handling & Services", "Additional Charges"}:
                subtotal = parse_float(row[2])
                if subtotal is not None:
                    subtotals[row[0]] = subtotal

        parsed_categories = [
            self._parse_storage_sheet(workbook["Storage Charges"], subtotals.get("Storage Charges")),
            self._parse_handling_sheet(workbook["Handling & Services"], subtotals.get("Handling & Services")),
            self._parse_additional_sheet(workbook["Additional Charges"], subtotals.get("Additional Charges")),
        ]
        categories: list[JsonObject] = [category for category in parsed_categories if category is not None]
        total_amount = (
            round(
                sum(
                    float(subtotal)
                    for category in categories
                    for subtotal in [category.get("subtotal")]
                    if isinstance(subtotal, (int, float))
                ),
                2,
            )
            if categories
            else None
        )

        return {
            "batch_number": batch_number,
            "invoice_date": invoice_date,
            "period_start": period_start,
            "period_end": period_end,
            "supplier_name": supplier_name,
            "client_name": client_name,
            "total_amount": total_amount,
            "currency": currency,
            "categories": categories,
            "extras": self._empty_extras(),
        }

    def _parse_storage_sheet(self, worksheet: object, subtotal: float | None) -> JsonObject | None:
        items: list[JsonObject] = []
        for row in worksheet.iter_rows(values_only=True):  # type: ignore[attr-defined]
            values = self._trim_row(list(row))
            if len(values) < 8 or parse_float(values[0]) is None:
                continue
            description = self._string_or_none(values[3])
            sku = self._string_or_none(values[2])
            if description is not None and sku is not None:
                description = f"{description} ({sku})"
            items.append(
                {
                    "description": description,
                    "amount": parse_float(values[7]),
                    "date": parse_date(self._string_or_none(values[1])),
                    "quantity": self._coerce_whole_number(parse_float(values[4])),
                    "unit_rate": parse_float(values[5]),
                    "reference": None,
                }
            )
        if not items:
            return None
        return {
            "category": "Storage Charges",
            "subtotal": subtotal if subtotal is not None else round(self._sum_amounts(items), 2),
            "line_items": items,
        }

    def _parse_handling_sheet(self, worksheet: object, subtotal: float | None) -> JsonObject | None:
        items: list[JsonObject] = []
        for row in worksheet.iter_rows(values_only=True):  # type: ignore[attr-defined]
            values = self._trim_row(list(row))
            if len(values) < 6 or parse_date(self._string_or_none(values[0])) is None:
                continue
            items.append(
                {
                    "description": self._string_or_none(values[1]),
                    "amount": parse_float(values[5]),
                    "date": parse_date(self._string_or_none(values[0])),
                    "quantity": self._coerce_whole_number(parse_float(values[3])),
                    "unit_rate": parse_float(values[4]),
                    "reference": self._string_or_none(values[2]),
                }
            )
        if not items:
            return None
        return {
            "category": "Handling & Services",
            "subtotal": subtotal if subtotal is not None else round(self._sum_amounts(items), 2),
            "line_items": items,
        }

    def _parse_additional_sheet(self, worksheet: object, subtotal: float | None) -> JsonObject | None:
        items: list[JsonObject] = []
        for row in worksheet.iter_rows(values_only=True):  # type: ignore[attr-defined]
            values = self._trim_row(list(row))
            if len(values) < 3:
                continue
            description = self._string_or_none(values[0])
            amount = parse_float(values[2])
            if description is None or amount is None or description.lower() == "description":
                continue
            if description.upper().startswith(("SUBTOTAL", "GRAND TOTAL")):
                continue
            items.append(
                {
                    "date": None,
                    "description": description,
                    "quantity": None,
                    "unit_rate": None,
                    "amount": amount,
                    "reference": self._string_or_none(values[1]),
                }
            )
        if not items:
            return None
        return {
            "category": "Additional Charges",
            "subtotal": subtotal if subtotal is not None else round(self._sum_amounts(items), 2),
            "line_items": items,
        }

    def _parse_period_range(self, line: str) -> tuple[str | None, str | None]:
        value = find_label_value(line, ["Billing Period"])
        if not value:
            return None, None
        parts = [segment.strip() for segment in value.split(" - ", maxsplit=1)]
        if len(parts) != 2:
            return None, None
        return parse_date(parts[0]), parse_date(parts[1])

    def _row_values(self, row: tuple[object, ...]) -> list[str]:
        values: list[str] = []
        for cell in row:
            string_value = self._string_or_none(cell)
            if string_value is not None:
                values.append(string_value)
        return values

    def _trim_row(self, values: list[object]) -> list[object]:
        trimmed = list(values)
        while trimmed and trimmed[0] in (None, ""):
            trimmed.pop(0)
        while trimmed and trimmed[-1] in (None, ""):
            trimmed.pop()
        return trimmed

    def _string_or_none(self, value: object) -> str | None:
        if value in (None, ""):
            return None
        return str(value).strip() or None

    def _parse_categories(self, text: str) -> list[JsonObject]:
        return []

    def _coerce_whole_number(self, value: float | None) -> int | float | None:
        if value is None:
            return None
        if value.is_integer():
            return int(value)
        return value

    def _empty_extras(self) -> dict[str, object]:
        return {
            "raw_labels": [],
            "unmapped_fields": {},
            "notes": [],
        }

    def _sum_amounts(self, items: list[JsonObject]) -> float:
        return sum(
            float(amount)
            for item in items
            for amount in [item.get("amount")]
            if isinstance(amount, (int, float))
        )
